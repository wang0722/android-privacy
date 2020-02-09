import openpyxl
import matplotlib.pyplot as plt
import numpy as np
from openpyxl import load_workbook
wb = load_workbook('元数据表.xlsx')
sheet=wb['Sheet1']
max_row=sheet.max_row
test_case=[]
aa=[]
bb=[]
for row in range(2,max_row+1):
    sub_data={}
    sub_data['序号']=sheet.cell(row,1).value
    sub_data['app名称']=sheet.cell(row,2).value
    sub_data['分类']=sheet.cell(row,3).value
    sub_data['下载量']=sheet.cell(row,4).value
    sub_data['好评率']=sheet.cell(row,5).value
    sub_data['评论数']=sheet.cell(row,6).value
    sub_data['星级'] = sheet.cell(row, 7).value
    sub_data['权限数量'] = sheet.cell(row, 8).value
    test_case.append(sub_data)
    aa.append(sub_data['下载量'])
    bb.append(sub_data['星级'])

# result=[]
# for i in aa:
#     try:
#         aaa=float(i)
#         result.append(aaa)
#     except ValueError:
#         aaa = i[:-1]
#         a = '亿'
#         b = '万'
#         if (a in i):
#             result.append(float(aaa) * 100000000)
#         elif (b in i):
#             result.append(float(aaa) * 10000)
#         else:
#             result.append(float(aaa))
# star=[]
# for i in bb:
#     bbb = i[:-1]
#     star.append(float(bbb)/20)
# row=2
# for i in result:
#     sheet.cell(row,4,i)
#     row=row+1
# row2=2
# for i in star:
#     sheet.cell(row2,7,i)
#     row2 = row2 + 1
# wb.save('元数据表.xlsx')

# #求均值
arr_mean = np.mean(aa)
#求方差
arr_var = np.var(aa)
#求标准差
arr_std = np.std(aa,ddof=1)
print("平均值为：%f" % arr_mean)
print("方差为：%f" % arr_var)
print("标准差为:%f" % arr_std)
z=[]
for i in aa:
    z.append((i - arr_mean) / arr_std)
print(z)
# z2=[]
# for i in z:
#     z2.append((i-arr_mean)/arr_std)

# print(z2)
# print(np.mean(z2))

# print(z2)
# plt.scatter(z2)
# plt.show()
# for x in aa:
#     x = float(x - np.min(aa))/(np.max(aa)- np.min(aa))
#     print(x)

for i in aa:
    c=(i-arr_mean)/arr_std
    if c>1:
        pass
    elif c< -1:
      pass
    else:
        z.append(c)
    z.append((i-arr_mean)/arr_std)
# print()