import time
import re
import requests
import os
import xlsxwriter
import subprocess
import openpyxl
import urllib.request
from bs4 import BeautifulSoup
import socket
from selenium import webdriver
import xlwt as ExcelWrite
socket.setdefaulttimeout(50)
browser = webdriver.Chrome()
url = 'https://www.wandoujia.com/top/app'
browser.get(url)
for i in range(1,50):
    browser.execute_script('window.scrollTo(0, document.body.scrollHeight)')
    time.sleep(0.2)
    try:
        browser.find_element_by_xpath('//*[@id="j-refresh-btn"]').click()
        time.sleep(0.2)
    except BaseException:
        pass
req11 = 'class="app-title-h2"><a href="(.*?)" title=".*?" class="name">.*?</a>'
aa = re.findall(req11, browser.page_source)
list = []
for i in aa:
    a = i.split('/')[-1]
    list.append(a)
browser.quit()  # 关闭并退出浏览器
# xls = ExcelWrite.Workbook(encoding='utf-8')
# xls2 = ExcelWrite.Workbook(encoding='utf-8')
# xls3 = ExcelWrite.Workbook(encoding='utf-8')
# sheet = xls.add_sheet("Sheet1")
# sheet2 = xls2.add_sheet("Sheet1")
# sheet3 = xls3.add_sheet("Sheet1")
# workbook2 = xlsxwriter.Workbook('元数据表.xlsx')     #创建工作簿
# sheet3 = workbook2.add_worksheet()            #创建工作表
# workbook1 = xlsxwriter.Workbook('实际权限permission.xlsx')     #创建工作簿
# sheet2 = workbook1.add_worksheet()            #创建工作表
# workbook = xlsxwriter.Workbook('权限信息result.xlsx')     #创建工作簿
# sheet = workbook.add_worksheet()            #创建工作表
data = openpyxl.Workbook() # 新建工作簿
data.create_sheet('Sheet1') # 添加页
sheet3 = data.active # 获得当前活跃的工作页，默认为第一个工作页
data1 = openpyxl.Workbook() # 新建工作簿
data1.create_sheet('Sheet1') # 添加页
sheet = data1.active # 获得当前活跃的工作页，默认为第一个工作页
data2 = openpyxl.Workbook() # 新建工作簿
data2.create_sheet('Sheet1') # 添加页
sheet2 = data2.active # 获得当前活跃的工作页，默认为第一个工作页
sheet.cell(1, 1, '序号')
sheet.cell(2, 1, 'app名称')
# sheet.write(2, 0, '应用版本')
sheet.cell(3, 1, '安卓要求')
sheet.cell(4, 1, '权限')
sheet2.cell(1, 1, '序号')
sheet2.cell(2, 1, 'app名称')
sheet2.cell(3, 1, '权限')
sheet3.cell(1,1 , '序号')
sheet3.cell(2, 1, 'app名称')
sheet3.cell(3, 1, 'app分类')
sheet3.cell(4, 1, '下载次数')
sheet3.cell(5, 1, '好评率')
sheet3.cell(6, 1, '评论数')
sheet3.cell(7, 1, '星级')
sheet3.cell(8, 1, '权限数量')
a = 2
c = 1
e = 2
f = 1
m = 2
n = 1
# 循环列表  爬取信息
number=1
print(len(list))
for ii in list:
    url12 = 'https://www.wandoujia.com/apps/{}'.format(ii)
    # url12 = 'https://www.wandoujia.com/apps/7925445'
    print(number)
    header = {
        'Accept': '*/*',
        'Accept-Encoding': 'gzip,deflate,sdch',
        'Accept-Language': 'zh-CN,zh;q=0.8,gl;q=0.6,zh-TW;q=0.4',
        'Connection': 'keep-alive',
        'Content-Type': 'application/x-www-form-urlencoded',
        'Referer': '{}'.format(url12),
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; WOW64) '
        'AppleWebKit/537.36 (KHTML, like Gecko) Chrome/63.0.3239.132 Safari/537.36'}
    try :
        html5 = requests.get(url12, headers=header, timeout=50).text
        req12 = 'itemprop="interactionCount" content="UserDownloads:.*?">(.*?)</i><b>次下载</b>'
        req13 = '<span>(.*?)好评率</span>'
        req14 = 'href="#comments" rel="nofollow"><i>(.*?)</i><b>人评论</b></a></div>'
        req15 = '<span class="avg-score-star"><i class="avg-score-current" style="width: (.*?)"></i>'
        req1 = '<div class="app-info"><p class="app-name"><span class="title" itemprop="name">(.*?)</span>'
        req2 = 'class="normal-dl-btn " data-bd-track="detail-common_download_main" href="(.*?)"'
        req111 = 'data-track="detail-click-appTag">(.*?)</a>'
        tag = re.findall(req111, html5)
        tag3 = []
        for i in tag:
            tag2 = i + ' '
            tag3.append(tag2)
        down = re.findall(req12, html5)
        percent1 = re.findall(req13, html5)
        if percent1 == []:
            percent = '暂无'
        else:
            percent = percent1
        comment = re.findall(req14, html5)
        star = re.findall(req15, html5)
        NAME = re.findall(req1, html5)
        download = re.findall(req2, html5)
        # print(NAME)
        # print(download)
        # print(down + percent + comment + star)
        sheet3.cell(1, m, n)
        sheet3.cell(2, m, NAME[0])

        sheet3.cell(3, m, tag3[0])
        sheet3.cell(4, m, down[0])
        sheet3.cell(5, m, percent[0])
        sheet3.cell(6, m, comment[0])
        sheet3.cell(7, m, star[0])

        n = n + 1

        req3 = 'itemprop="permissions">(.*?)</span></li>'
        permission = re.findall(req3, html5)
        req6 = 'itemprop="operatingSystems" content="Android">(.*?)<div>'
        req7 = 'data-app-vname="version (.*?)"'
        version = re.findall(req7, html5)
        try:
            android = re.findall(req6, html5)[0]
        except IndexError:
            android = '未知'
        sheet.cell(1, a, c)
        sheet.cell(2, a, NAME[0])
        # sheet.write(2, a, version)
        sheet.cell(3, a, android)
        num = 4
        for d in permission:
            sheet.cell(num, a, d)
            num = num + 1
        a = a + 1
        c = c + 1
        if version == []:
            version1 = ['最新版']
        else:
            version1 = version
        try:
            if os.path.exists(
                    'app\\' + '{}'.format(NAME[0].replace(' ', '')) + '{}.apk'.format(version1[0].replace(' ', ''))):
                command = 'aapt dump badging E:\\pythontest\\杂乱\\测试二号\\app\\{}'.format(NAME[0].replace(
                    ' ', '')) + '{}.apk'.format(version1[0].replace(' ', ''))  # 可以直接在命令行中执行的命令
                output = subprocess.Popen(
                    command,
                    stdout=subprocess.PIPE,
                    shell=True)  # .communicate()[0]
                aa = output.stdout.read()
                try:
                    result = str(aa, 'utf-8')
                except UnicodeDecodeError:
                    result = str(aa)
                req = "uses-permission:'(.*?)'"
                res = re.findall(req, result)
                sheet2.cell(1, e, f)
                sheet2.cell(2, e, NAME[0])
                num1 = 3
                for dd in res:
                    sheet2.cell(num1, e, dd)
                    num1 = num1 + 1
                sheet3.cell(8, m, len(res))
                e = e + 1
                f = f + 1
                m = m + 1
                number = number + 1
                data.save('元数据表.xlsx')
                data1.save('权限信息result.xlsx')
                data2.save('实际权限permission.xlsx')
                pass
            else:
                try:
                    urllib.request.urlretrieve(download[0],
                                               'app\\' + '{}'.format(NAME[0].replace(' ', '')) + '{}.apk'.format(
                                                   version1[0].replace(' ', '')))
                    command = 'aapt dump badging E:\\pythontest\\杂乱\\测试二号\\app\\{}'.format(NAME[0].replace(
                        ' ', '')) + '{}.apk'.format(version1[0].replace(' ', ''))  # 可以直接在命令行中执行的命令
                    output = subprocess.Popen(
                        command,
                        stdout=subprocess.PIPE,
                        shell=True)  # .communicate()[0]
                    aa = output.stdout.read()
                    try:
                        result = str(aa, 'utf-8')
                    except UnicodeDecodeError:
                        result = str(aa)
                    req = "uses-permission:'(.*?)'"
                    res = re.findall(req, result)
                    sheet2.cell(1, e, f)
                    sheet2.cell(2, e, NAME[0])
                    num1 = 3
                    for dd in res:
                        sheet2.cell(num1, e, dd)
                        num1 = num1 + 1
                    sheet3.cell(8, m, len(res))
                    e = e + 1
                    f = f + 1
                    m = m + 1
                    number = number + 1
                    data.save('元数据表.xlsx')
                    data1.save('权限信息result.xlsx')
                    data2.save('实际权限permission.xlsx')
                except:
                    data.save('元数据表.xlsx')
                    data1.save('权限信息result.xlsx')
                    data2.save('实际权限permission.xlsx')
                    continue
        except:
            data.save('元数据表.xlsx')
            data1.save('权限信息result.xlsx')
            data2.save('实际权限permission.xlsx')
            continue
    except:
        data.save('元数据表.xlsx')
        data1.save('权限信息result.xlsx')
        data2.save('实际权限permission.xlsx')
        continue
data.save('元数据表.xlsx')
data1.save('权限信息result.xlsx')
data2.save('实际权限permission.xlsx')