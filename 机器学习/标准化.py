from sklearn.preprocessing import StandardScaler
import openpyxl
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from openpyxl import load_workbook
wb = load_workbook('元数据表.xlsx')
sheet=wb['Sheet1']
max_row=sheet.max_row
test_case=[]
aa=[]
bb=[]
for row in range(2,max_row+1):
    sub_data={}
    sub_data['序号']=sheet.cell(row,1).value
    sub_data['app名称']=sheet.cell(row,2).value
    sub_data['分类']=sheet.cell(row,3).value
    sub_data['下载量']=sheet.cell(row,4).value
    sub_data['好评率']=sheet.cell(row,5).value
    sub_data['评论数']=sheet.cell(row,6).value
    sub_data['星级'] = sheet.cell(row, 7).value
    sub_data['权限数量'] = sheet.cell(row, 8).value
    test_case.append(sub_data)
    aa.append(sub_data['下载量'])
    bb.append(sub_data['星级'])
data=pd.read_excel("元数据表.xlsx")
data=data.iloc[:,[3]]
# print(data1)
tran=StandardScaler()
dataa=tran.fit_transform(data)
# x_train = pd.DataFrame(dataa)
c=tran.fit_transform(data)
print("data new:",c)